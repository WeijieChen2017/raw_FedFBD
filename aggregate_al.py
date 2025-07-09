results_dir = "fbd_run"

import os
import glob
import openpyxl
import json
import numpy as np
folder_list = glob.glob(os.path.join(results_dir, "*_al*"))

# write the header to the excel file
# create  the xlsx file if it does not exist
if not os.path.exists('al.xlsx'):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = "dataset"
    sheet.cell(row=1, column=2).value = "metrics"
    sheet.cell(row=1, column=3).value = "len_rounds"

    workbook.save('al.xlsx')

workbook = openpyxl.load_workbook('al.xlsx')
sheet = workbook.active

for i_row, folder in enumerate(folder_list):
    if "_al" in folder:
        print(folder)
        folder_name = os.path.basename(folder)
        segments = folder_name.split("_")
        
        metrics = segments[-1]

        dataset = segments[0]

        # write the tau value to the excel file
        sheet.cell(row=i_row+2, column=1).value = dataset
        sheet.cell(row=i_row+2, column=2).value = metrics

        train_results = os.path.join(folder, "eval_results", "server_evaluation_history.json")
        try:
            with open(train_results, "r") as f:
                train_results = json.load(f)
            
            len_rounds = len(train_results)
            sheet.cell(row=i_row+2, column=4).value = len_rounds
            for i_round, round_results in enumerate(train_results):
                start_idx = 5 + i_round * 7
                sheet.cell(row=i_row+2, column=start_idx+i_round).value = round_results["round"]
                M_auc = []
                M_acc = []
                for i_M in range(6):
                    M_auc.append(round_results[f"M{i_M}"]["test_auc"])
                    M_acc.append(round_results[f"M{i_M}"]["test_acc"])

                sheet.cell(row=i_row+2, column=start_idx+i_round+1).value = np.mean(M_auc)
                sheet.cell(row=i_row+2, column=start_idx+i_round+2).value = np.mean(M_acc)
                sheet.cell(row=i_row+2, column=start_idx+i_round+3).value = round_results[f"averaging"]["test_auc"]
                sheet.cell(row=i_row+2, column=start_idx+i_round+4).value = round_results[f"averaging"]["test_acc"]
                sheet.cell(row=i_row+2, column=start_idx+i_round+5).value = round_results[f"ensemble"]["test_auc"]
                sheet.cell(row=i_row+2, column=start_idx+i_round+6).value = round_results[f"ensemble"]["test_acc"]
        except:
            print(f"No train results found for {folder}")
            continue


workbook.save('al.xlsx')
print(f"Saved {i_row+1} rows to al.xlsx")